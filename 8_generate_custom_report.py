# 8_generate_custom_report.py
# 权威最终版: 应用了您最终确认的、包含所有层级和相对位置的定制化布局。

import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# --- 【【【每次运行时，请根据您的项目情况修改此处的配置】】】 ---

# 1. 报告头部信息
PROJECT_INFO = {
    "项目号": "Project-2025-Q3",
    "题号": "Q5, Q8",
    "问题": "您对本次入住的整体体验如何？, 您有什么改进建议吗？"
}

# 2. 最终审核、修正后的编码结果文件名 (聚合格式)
CORRECTED_DATA_FILE = "07_final_reports/Final_Coded_Report_CORRECTED_AGGREGATED.csv" 

# 3. 您上一期的原始基础码表
BASE_CODEBOOK_FILE = "01_source_data/last_phase_codebook.csv"

# 4. 您为新题目手动定稿的码表 (如果本次项目没有新题，此文件可以不存在)
NEW_QUESTION_CODEBOOK_FILE = "new_question_codebook.csv" 

# 5. 最终输出的Excel文件名
OUTPUT_EXCEL_FILE = "07_final_reports/Custom_Codebook_Report.xlsx"
# -----------------------------------------------------------

def generate_custom_report():
    """
    主函数，生成高度定制化的Excel码表报告。
    """
    print("--- 开始生成高度定制化的最终码表Excel文件 ---")
    
    try:
        corrected_df = pd.read_csv(CORRECTED_DATA_FILE, dtype=str).fillna('')
        base_codebook_df = pd.read_csv(BASE_CODEBOOK_FILE, dtype=str).fillna('')
    except FileNotFoundError as e:
        print(f"错误：找不到必需的文件 {e.filename}。请检查配置。")
        return

    # --- 1. 提取本次编码实际使用过的所有唯一code ---
    if 'code_agg' not in corrected_df.columns:
        print(f"错误：修正后的报告中缺少 'code_agg' 列。")
        return
        
    unique_codes_used = set()
    for codes in corrected_df['code_agg'].dropna():
        unique_codes_used.update(str(codes).split('; '))
    unique_codes_used.discard('')
    print(f"在本次编码中共计使用了 {len(unique_codes_used)} 个唯一编码。")

    # --- 2. 构建包含所有可用编码的“超级码表” ---
    try:
        new_question_codebook_df = pd.read_csv(NEW_QUESTION_CODEBOOK_FILE, dtype=str).fillna('')
    except FileNotFoundError:
        new_question_codebook_df = pd.DataFrame()
        
    super_codebook_df = pd.concat([base_codebook_df, new_question_codebook_df], ignore_index=True).drop_duplicates(subset=['code'])
    
    # --- 3. 按需筛选，得到最终的项目专属码表 ---
    final_codebook_df = super_codebook_df[super_codebook_df['code'].isin(unique_codes_used)].copy()
    
    # --- 4. 数据准备，为排序和分组做准备 ---
    # 填充空值以便进行后续的分组和排序
    final_codebook_df['sentiment'] = final_codebook_df['sentiment'].replace('', 'z_no_sentiment')
    final_codebook_df['net'] = final_codebook_df['net'].replace('', 'z_no_net')
    final_codebook_df['subnet'] = final_codebook_df['subnet'].replace('', 'z_no_subnet')
    
    # 按层级排序
    final_codebook_df.sort_values(by=['sentiment', 'net', 'subnet', 'code'], inplace=True)
    
    # 分离出“有分类”和“无分类”的编码
    categorized_df = final_codebook_df[
        (final_codebook_df['sentiment'] != 'z_no_sentiment') |
        (final_codebook_df['net'] != 'z_no_net') |
        (final_codebook_df['subnet'] != 'z_no_subnet')
    ]
    uncategorized_df = final_codebook_df[
        (final_codebook_df['sentiment'] == 'z_no_sentiment') &
        (final_codebook_df['net'] == 'z_no_net') &
        (final_codebook_df['subnet'] == 'z_no_subnet')
    ]

    # --- 5. 使用 openpyxl 创建并写入Excel ---
    print("正在创建和写入Excel文件，请稍候...")
    wb = Workbook()
    ws = wb.active
    ws.title = "项目码表"
    
    # 定义样式
    bold_font = Font(bold=True)
    
    # 写入文件头部
    ws['A1'] = "项目号:"
    ws['B1'] = PROJECT_INFO.get("项目号", "")
    
    ws['A2'] = f"题号:{PROJECT_INFO.get('题号', '')}"
    
    ws['A3'] = f"问题:{PROJECT_INFO.get('问题', '')}"
    
    current_row = 6 # 从第6行开始写入数据（空两行）

    # 写入第二部分：有分类的编码区域
    last_sentiment, last_net, last_subnet = None, None, None
    for index, row_data in categorized_df.iterrows():
        # 恢复原始的空值用于判断和显示
        sentiment = row_data['sentiment'] if row_data['sentiment'] != 'z_no_sentiment' else ''
        net = row_data['net'] if row_data['net'] != 'z_no_net' else ''
        subnet = row_data['subnet'] if row_data['subnet'] != 'z_no_subnet' else ''
        
        # 写入最高层级 - Sentiment
        if sentiment and sentiment != last_sentiment:
            ws[f'A{current_row}'] = f"sentiment {sentiment}"
            ws[f'A{current_row}'].font = Font(bold=True, color="FF6347") # 橙红色
            current_row += 1
            last_sentiment = sentiment
            last_net, last_subnet = None, None # 重置下级
            
        # 写入次高层级 - Net
        if net and net != last_net:
            ws[f'A{current_row}'] = f"net {net}"
            ws[f'A{current_row}'].font = Font(bold=True, color="4682B4") # 钢蓝色
            current_row += 1
            last_net = net
            last_subnet = None # 重置下级

        # --- 【【【核心布局修正处】】】 ---
        if subnet and subnet != last_subnet:
            ws[f'B{current_row}'] = f"subnet {subnet}"
            ws[f'B{current_row}'].font = Font(bold=True, color="32CD32") # 酸橙绿
            current_row += 1
            last_subnet = subnet
            
            # 有subnet的code，现在写入B/C列
            ws[f'B{current_row}'] = row_data['code']
            ws[f'C{current_row}'] = row_data['label']
            current_row += 1

        elif net and not subnet: # 有net，但没有subnet
            ws[f'B{current_row}'] = row_data['code']
            ws[f'C{current_row}'] = row_data['label']
            current_row += 1
            
        elif subnet: # 同一个subnet下的其他code
            ws[f'B{current_row}'] = row_data['code']
            ws[f'C{current_row}'] = row_data['label']
            current_row += 1
        
        elif sentiment and not net and not subnet: # 只有sentiment
            ws[f'B{current_row}'] = row_data['code']
            ws[f'C{current_row}'] = row_data['label']
            current_row += 1
        # ------------------------------------

    # 写入第三部分：无分类的编码区域
    if not uncategorized_df.empty:
        current_row += 2 # 中间空两行
        ws[f'A{current_row}'] = "无分类编码"
        ws[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
        
        for index, row_data in uncategorized_df.iterrows():
            ws[f'A{current_row}'] = row_data['code']
            ws[f'B{current_row}'] = row_data['label']
            current_row += 1

    # 调整列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20

    # 保存文件
    wb.save(OUTPUT_EXCEL_FILE)
    print("-" * 50)
    print(f"成功生成最终定制化码表报告: {OUTPUT_EXCEL_FILE}")
    print("-" * 50)


if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("错误：缺少 openpyxl 库。请运行: pip install openpyxl")
        exit()
    generate_custom_report()